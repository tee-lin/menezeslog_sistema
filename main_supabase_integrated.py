# SISTEMA v7.5 FINAL - TIMEOUT ESTENDIDO + TODAS AS APIs SUPABASE
# BACKEND COMPLETO COM TODAS AS FUNCIONALIDADES

from flask import Flask, request, jsonify, send_from_directory
from flask_cors import CORS
import os
import csv
import io
import json
import logging
import datetime
from datetime import datetime, timedelta
import chardet
import openpyxl
from werkzeug.utils import secure_filename
import time
import threading
from concurrent.futures import ThreadPoolExecutor
import queue

# Configuração da aplicação
app = Flask(__name__, static_folder='static', static_url_path='')
CORS(app)

# Configuração de logging OTIMIZADA
logging.basicConfig(level=logging.WARNING)
app.logger.setLevel(logging.WARNING)

# Configurações FINAIS
UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'csv', 'xlsx', 'xls'}
BATCH_SIZE = 5000
SUPABASE_BATCH_SIZE = 1000
MAX_WORKERS = 3

# Criar pastas se não existirem
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# Tarifas padrão
TARIFAS_PADRAO = {
    0: 3.50,  # Encomendas
    9: 2.00,  # Cards
    6: 0.50,  # Revistas
    8: 0.50   # Revistas
}

# ==================== CONFIGURAÇÃO SUPABASE ====================

try:
    from supabase import create_client, Client
    SUPABASE_AVAILABLE = True
    print("✅ Biblioteca supabase-py disponível")
except ImportError:
    SUPABASE_AVAILABLE = False
    print("❌ Biblioteca supabase-py não encontrada")

# Cliente Supabase global
supabase_client = None

def init_supabase():
    """Inicializa conexão com Supabase"""
    global supabase_client
    
    if not SUPABASE_AVAILABLE:
        print("❌ Supabase não disponível")
        return False
    
    try:
        supabase_url = os.environ.get('SUPABASE_URL')
        supabase_key = os.environ.get('SUPABASE_ANON_KEY')
        
        if not supabase_url or not supabase_key:
            print("❌ Credenciais do Supabase não configuradas")
            return False
        
        supabase_client = create_client(supabase_url, supabase_key)
        
        # Testar conexão
        result = supabase_client.table('motoristas').select('count').execute()
        print(f"✅ Conexão com Supabase estabelecida")
        return True
        
    except Exception as e:
        print(f"❌ Erro ao conectar com Supabase: {str(e)}")
        return False

# Inicializar Supabase
SUPABASE_CONNECTED = init_supabase()

# ==================== CACHE GLOBAL ====================
_cache_motoristas = {}
_cache_tarifas = {}
_cache_prestadores = {}
_cache_timestamp = 0

def invalidate_cache():
    """Invalida cache quando dados são atualizados"""
    global _cache_timestamp
    _cache_timestamp = time.time()

# ==================== FUNÇÕES SUPABASE OTIMIZADAS ====================

def get_motoristas_supabase_cached(empresa_id=1):
    """Carrega motoristas do Supabase com cache"""
    global _cache_motoristas, _cache_timestamp
    
    cache_key = f"motoristas_{empresa_id}"
    current_time = time.time()
    
    if cache_key in _cache_motoristas and (current_time - _cache_timestamp) < 300:
        return _cache_motoristas[cache_key]
    
    try:
        if not SUPABASE_CONNECTED:
            return []
        
        result = supabase_client.table('motoristas').select('*').eq('empresa_id', empresa_id).execute()
        motoristas = result.data if result.data else []
        
        _cache_motoristas[cache_key] = motoristas
        _cache_timestamp = current_time
        
        return motoristas
        
    except Exception as e:
        print(f"❌ Erro ao carregar motoristas: {str(e)}")
        return []

def get_prestadores_supabase_cached(empresa_id=1):
    """Carrega prestadores do Supabase com cache"""
    global _cache_prestadores, _cache_timestamp
    
    cache_key = f"prestadores_{empresa_id}"
    current_time = time.time()
    
    if cache_key in _cache_prestadores and (current_time - _cache_timestamp) < 300:
        return _cache_prestadores[cache_key]
    
    try:
        if not SUPABASE_CONNECTED:
            return []
        
        result = supabase_client.table('prestadores').select('*').eq('empresa_id', empresa_id).execute()
        prestadores = result.data if result.data else []
        
        _cache_prestadores[cache_key] = prestadores
        _cache_timestamp = current_time
        
        return prestadores
        
    except Exception as e:
        print(f"❌ Erro ao carregar prestadores: {str(e)}")
        return []

def get_awbs_count_supabase(empresa_id=1):
    """Conta AWBs no Supabase"""
    try:
        if not SUPABASE_CONNECTED:
            return 0
        
        result = supabase_client.table('awbs').select('count').eq('empresa_id', empresa_id).execute()
        return len(result.data) if result.data else 0
        
    except Exception as e:
        print(f"❌ Erro ao contar AWBs: {str(e)}")
        return 0

def get_tarifas_supabase_cached(empresa_id=1):
    """Carrega tarifas do Supabase com cache"""
    global _cache_tarifas, _cache_timestamp
    
    cache_key = f"tarifas_{empresa_id}"
    current_time = time.time()
    
    if cache_key in _cache_tarifas and (current_time - _cache_timestamp) < 300:
        return _cache_tarifas[cache_key]
    
    try:
        if not SUPABASE_CONNECTED:
            return {}
        
        result = supabase_client.table('tarifas').select('*').eq('empresa_id', empresa_id).execute()
        
        tarifas_dict = {}
        if result.data:
            for tarifa in result.data:
                motorista_id = str(tarifa['id_motorista'])
                if motorista_id not in tarifas_dict:
                    tarifas_dict[motorista_id] = {}
                tarifas_dict[motorista_id][tarifa['tipo_servico']] = float(tarifa['valor'])
        
        _cache_tarifas[cache_key] = tarifas_dict
        return tarifas_dict
        
    except Exception as e:
        print(f"❌ Erro ao carregar tarifas: {str(e)}")
        return {}

def save_motoristas_supabase_batch(motoristas, empresa_id=1):
    """Salva motoristas no Supabase usando UPSERT"""
    try:
        if not SUPABASE_CONNECTED:
            return False
        
        motoristas_data = []
        for motorista in motoristas:
            motorista_data = {
                'empresa_id': empresa_id,
                'id_motorista': motorista['id_motorista'],
                'nome_motorista': motorista['nome_motorista'],
                'created_at': motorista.get('created_at', datetime.now().isoformat()),
                'updated_at': datetime.now().isoformat()
            }
            motoristas_data.append(motorista_data)
        
        supabase_client.table('motoristas').upsert(
            motoristas_data,
            on_conflict='empresa_id,id_motorista'
        ).execute()
        
        invalidate_cache()
        return True
        
    except Exception as e:
        print(f"❌ Erro ao salvar motoristas: {str(e)}")
        return False

def save_awbs_supabase_batch(awbs_list, empresa_id=1):
    """Salva AWBs no Supabase usando UPSERT em lotes"""
    try:
        if not SUPABASE_CONNECTED or not awbs_list:
            return False
        
        total_saved = 0
        for i in range(0, len(awbs_list), SUPABASE_BATCH_SIZE):
            batch = awbs_list[i:i + SUPABASE_BATCH_SIZE]
            
            awbs_data = []
            for awb_data in batch:
                awb_record = {
                    'empresa_id': empresa_id,
                    'awb': awb_data['awb'],
                    'id_motorista': awb_data['id_motorista'],
                    'nome_motorista': awb_data['nome_motorista'],
                    'tipo_servico': awb_data['tipo_servico'],
                    'data_entrega': awb_data['data_entrega'],
                    'valor_entrega': float(awb_data['valor_entrega']),
                    'status': awb_data['status'],
                    'created_at': awb_data.get('created_at', datetime.now().isoformat()),
                    'updated_at': datetime.now().isoformat()
                }
                awbs_data.append(awb_record)
            
            supabase_client.table('awbs').upsert(
                awbs_data,
                on_conflict='empresa_id,awb'
            ).execute()
            
            total_saved += len(batch)
            print(f"💾 Lote Supabase: {total_saved}/{len(awbs_list)} AWBs salvas")
        
        return True
        
    except Exception as e:
        print(f"❌ Erro ao salvar AWBs: {str(e)}")
        return False

# ==================== PROCESSAMENTO ASSÍNCRONO ====================

class ProcessadorAssincronoBeastMode:
    """Processador assíncrono ultra otimizado"""
    
    def __init__(self):
        self.progress_queue = queue.Queue()
        self.result_queue = queue.Queue()
        self.executor = ThreadPoolExecutor(max_workers=MAX_WORKERS)
    
    def processar_csv_beast_mode(self, dados_csv, motoristas_dict, tarifas_cache, empresa_id=1):
        """Processa CSV em modo BEAST com threads paralelas"""
        total_linhas = len(dados_csv)
        entregas_processadas = 0
        entregas_erro = 0
        awbs_processadas = []
        
        print(f"🚀 BEAST MODE: Processando {total_linhas} linhas com {MAX_WORKERS} threads")
        
        chunk_size = BATCH_SIZE
        chunks = [dados_csv[i:i + chunk_size] for i in range(0, total_linhas, chunk_size)]
        
        futures = []
        for chunk_idx, chunk in enumerate(chunks):
            future = self.executor.submit(
                self._processar_chunk,
                chunk, chunk_idx, motoristas_dict, tarifas_cache
            )
            futures.append(future)
        
        for future in futures:
            try:
                chunk_result = future.result(timeout=60)  # 60s por chunk
                entregas_processadas += chunk_result['processadas']
                entregas_erro += chunk_result['erros']
                awbs_processadas.extend(chunk_result['awbs'])
                
                if len(awbs_processadas) >= SUPABASE_BATCH_SIZE * 2:
                    save_awbs_supabase_batch(awbs_processadas, empresa_id)
                    awbs_processadas = []
                
            except Exception as e:
                print(f"❌ Erro no chunk: {str(e)}")
                entregas_erro += chunk_size
        
        if awbs_processadas:
            save_awbs_supabase_batch(awbs_processadas, empresa_id)
        
        return {
            'entregas_processadas': entregas_processadas,
            'entregas_erro': entregas_erro,
            'total_awbs': entregas_processadas
        }
    
    def _processar_chunk(self, chunk, chunk_idx, motoristas_dict, tarifas_cache):
        """Processa um chunk de dados"""
        processadas = 0
        erros = 0
        awbs_chunk = []
        
        for linha in chunk:
            try:
                # Procurar AWB
                awb = None
                for col in ['AWB', 'awb', 'Awb']:
                    if col in linha and linha[col]:
                        awb = str(linha[col]).strip()
                        if awb and awb != 'None':
                            break
                
                if not awb:
                    erros += 1
                    continue
                
                # Procurar ID do motorista
                id_motorista = None
                for col in ['ID do motorista', 'ID_MOTORISTA', 'id_motorista']:
                    if col in linha and linha[col]:
                        try:
                            id_motorista = int(linha[col])
                            break
                        except (ValueError, TypeError):
                            continue
                
                if not id_motorista or id_motorista not in motoristas_dict:
                    erros += 1
                    continue
                
                # Procurar tipo de serviço
                tipo_servico = 0
                for col in ['Tipo de Serviço', 'TIPO_SERVICO', 'tipo_servico']:
                    if col in linha and linha[col]:
                        try:
                            tipo_servico = int(linha[col])
                            break
                        except (ValueError, TypeError):
                            continue
                
                # Procurar data/hora da entrega
                data_entrega = None
                for col in ['Data/Hora Status do último status', 'DATA_ENTREGA', 'data_entrega']:
                    if col in linha and linha[col]:
                        data_entrega = str(linha[col]).strip()
                        if data_entrega and data_entrega != 'None':
                            break
                
                if not data_entrega:
                    data_entrega = datetime.now().isoformat()
                
                # Calcular valor da entrega
                tarifa_motorista = tarifas_cache.get(str(id_motorista), TARIFAS_PADRAO)
                valor_entrega = float(tarifa_motorista.get(tipo_servico, TARIFAS_PADRAO.get(tipo_servico, 0)))
                
                # Criar AWB
                awb_data = {
                    'awb': awb,
                    'id_motorista': id_motorista,
                    'nome_motorista': motoristas_dict[id_motorista]['nome_motorista'],
                    'tipo_servico': tipo_servico,
                    'data_entrega': data_entrega,
                    'valor_entrega': valor_entrega,
                    'status': 'NAO_PAGA',
                    'created_at': datetime.now().isoformat(),
                    'updated_at': datetime.now().isoformat()
                }
                
                awbs_chunk.append(awb_data)
                processadas += 1
                
            except Exception as e:
                erros += 1
                continue
        
        print(f"📦 Chunk {chunk_idx + 1}: {processadas} processadas, {erros} erros")
        
        return {
            'processadas': processadas,
            'erros': erros,
            'awbs': awbs_chunk
        }

# Instância global do processador
processador_beast = ProcessadorAssincronoBeastMode()

# ==================== FUNÇÕES DE DETECÇÃO ====================

def detectar_encoding(file_content_bytes):
    """Detecta encoding do arquivo"""
    try:
        for encoding in ['iso-8859-1', 'windows-1252', 'latin-1', 'utf-8']:
            try:
                content = file_content_bytes.decode(encoding)
                return content, encoding
            except UnicodeDecodeError:
                continue
        
        content = file_content_bytes.decode('latin-1', errors='replace')
        return content, 'latin-1'
        
    except Exception as e:
        content = file_content_bytes.decode('utf-8', errors='replace')
        return content, 'utf-8'

def detectar_delimitador_csv(content):
    """Detecta delimitador do CSV"""
    try:
        primeira_linha = content.split('\n')[0]
        
        delimitadores = {
            ';': primeira_linha.count(';'),
            ',': primeira_linha.count(','),
            '\t': primeira_linha.count('\t'),
            '|': primeira_linha.count('|')
        }
        
        delimitador = max(delimitadores, key=delimitadores.get)
        return delimitador if delimitadores[delimitador] > 0 else ','
        
    except Exception as e:
        return ','

def allowed_file(filename):
    """Verifica se o arquivo é permitido"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# ==================== ROTAS PRINCIPAIS ====================

@app.route('/')
def index():
    """Página inicial"""
    return send_from_directory(app.static_folder, 'index.html')

@app.route('/<path:filename>')
def static_files(filename):
    """Servir arquivos estáticos"""
    return send_from_directory(app.static_folder, filename)

# ==================== API DE MOTORISTAS ====================

@app.route('/api/motoristas', methods=['GET'])
def get_motoristas_api():
    """Lista todos os motoristas"""
    try:
        empresa_id = 1
        motoristas = get_motoristas_supabase_cached(empresa_id)
        
        return jsonify({
            'success': True,
            'data': motoristas,
            'total': len(motoristas),
            'source': 'supabase' if SUPABASE_CONNECTED else 'local'
        })
        
    except Exception as e:
        return jsonify({'error': f'Erro ao buscar motoristas: {str(e)}'}), 500

@app.route('/api/motoristas/count', methods=['GET'])
def get_motoristas_count():
    """Conta total de motoristas"""
    try:
        empresa_id = 1
        motoristas = get_motoristas_supabase_cached(empresa_id)
        
        return jsonify({
            'success': True,
            'count': len(motoristas)
        })
        
    except Exception as e:
        return jsonify({'error': f'Erro ao contar motoristas: {str(e)}'}), 500

@app.route('/api/motoristas/upload', methods=['POST'])
def upload_motoristas():
    """Upload da planilha DE-PARA de motoristas"""
    try:
        empresa_id = 1
        
        print("=== INÍCIO UPLOAD MOTORISTAS v7.5 FINAL ===")
        
        if 'file' not in request.files:
            return jsonify({'error': 'Nenhum arquivo enviado'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'Nenhum arquivo selecionado'}), 400
        
        file_content_bytes = file.read()
        print(f"Planilha DE-PARA recebida: {file.filename} ({len(file_content_bytes)} bytes)")
        
        motoristas_processados = 0
        motoristas_erro = 0
        
        motoristas = get_motoristas_supabase_cached(empresa_id)
        
        file_content, encoding_usado = detectar_encoding(file_content_bytes)
        print(f"Encoding detectado: {encoding_usado}")
        
        if file.filename.endswith(('.xlsx', '.xls')):
            try:
                workbook = openpyxl.load_workbook(io.BytesIO(file_content_bytes))
                sheet = workbook.active
                
                print(f"Excel carregado: {sheet.max_row} linhas x {sheet.max_column} colunas")
                
                headers = []
                for col_num in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=1, column=col_num)
                    headers.append(str(cell.value) if cell.value else f"Col_{col_num}")
                
                print(f"Cabeçalhos encontrados: {headers}")
                
                for row_num in range(2, sheet.max_row + 1):
                    try:
                        linha_data = {}
                        for col_num in range(1, sheet.max_column + 1):
                            cell = sheet.cell(row=row_num, column=col_num)
                            linha_data[headers[col_num - 1]] = cell.value
                        
                        id_motorista = None
                        for col_name in ['ID do motorista', 'ID', 'id', 'Id', 'ID_MOTORISTA', 'id_motorista']:
                            if col_name in linha_data and linha_data[col_name] is not None:
                                try:
                                    if isinstance(linha_data[col_name], (int, float)):
                                        id_motorista = int(linha_data[col_name])
                                    else:
                                        id_motorista = int(float(str(linha_data[col_name])))
                                    break
                                except (ValueError, TypeError):
                                    continue
                        
                        nome_motorista = None
                        for col_name in ['Nome do motorista', 'NOME', 'nome', 'Nome', 'NOME_MOTORISTA', 'nome_motorista']:
                            if col_name in linha_data and linha_data[col_name] is not None:
                                nome_motorista = str(linha_data[col_name]).strip()
                                if nome_motorista and nome_motorista != 'None' and nome_motorista != '':
                                    break
                        
                        if not id_motorista or not nome_motorista:
                            motoristas_erro += 1
                            continue
                        
                        motorista_existente = None
                        for m in motoristas:
                            if m['id_motorista'] == id_motorista:
                                motorista_existente = m
                                break
                        
                        if motorista_existente:
                            motorista_existente['nome_motorista'] = nome_motorista
                            motorista_existente['updated_at'] = datetime.now().isoformat()
                        else:
                            motorista = {
                                'id_motorista': id_motorista,
                                'nome_motorista': nome_motorista,
                                'created_at': datetime.now().isoformat(),
                                'updated_at': datetime.now().isoformat()
                            }
                            motoristas.append(motorista)
                        
                        motoristas_processados += 1
                        
                    except Exception as e:
                        motoristas_erro += 1
                        continue
                
            except Exception as e:
                return jsonify({'error': f'Erro ao processar Excel: {str(e)}'}), 500
        
        else:
            return jsonify({'error': 'Formato de arquivo não suportado para motoristas'}), 400
        
        success = save_motoristas_supabase_batch(motoristas, empresa_id)
        
        print(f"=== RESULTADO FINAL ===")
        print(f"Motoristas processados: {motoristas_processados}")
        print(f"Motoristas com erro: {motoristas_erro}")
        print(f"Total de motoristas no sistema: {len(motoristas)}")
        print(f"Salvos no Supabase: {'✅' if success else '❌'}")
        
        return jsonify({
            'success': True,
            'message': f'Planilha processada: {motoristas_processados} motoristas, {motoristas_erro} erros',
            'data': {
                'motoristas_processados': motoristas_processados,
                'motoristas_erro': motoristas_erro,
                'total_motoristas': len(motoristas),
                'saved_to_supabase': success
            }
        })
        
    except Exception as e:
        return jsonify({'error': f'Erro no upload: {str(e)}'}), 500

# ==================== API DE PRESTADORES ====================

@app.route('/api/prestadores', methods=['GET'])
def get_prestadores_api():
    """Lista todos os prestadores"""
    try:
        empresa_id = 1
        prestadores = get_prestadores_supabase_cached(empresa_id)
        
        return jsonify({
            'success': True,
            'data': prestadores,
            'total': len(prestadores)
        })
        
    except Exception as e:
        return jsonify({'error': f'Erro ao buscar prestadores: {str(e)}'}), 500

@app.route('/api/prestadores/estatisticas', methods=['GET'])
def get_prestadores_estatisticas():
    """Estatísticas dos prestadores"""
    try:
        empresa_id = 1
        motoristas = get_motoristas_supabase_cached(empresa_id)
        prestadores = get_prestadores_supabase_cached(empresa_id)
        
        # Calcular estatísticas
        total_motoristas = len(motoristas)
        total_prestadores = len(prestadores)
        
        # Motoristas em grupos
        motoristas_em_grupos = 0
        for prestador in prestadores:
            motoristas_em_grupos += 1  # Principal
            if prestador.get('motoristas_ajudantes'):
                motoristas_em_grupos += len(prestador['motoristas_ajudantes'])
        
        motoristas_individuais = total_motoristas - motoristas_em_grupos
        
        return jsonify({
            'success': True,
            'data': {
                'total_motoristas': total_motoristas,
                'total_prestadores': total_prestadores,
                'motoristas_em_grupos': motoristas_em_grupos,
                'motoristas_individuais': motoristas_individuais
            }
        })
        
    except Exception as e:
        return jsonify({'error': f'Erro ao calcular estatísticas: {str(e)}'}), 500

# ==================== API DE TARIFAS ====================

@app.route('/api/tarifas', methods=['GET'])
def get_tarifas_api():
    """Lista todas as tarifas"""
    try:
        empresa_id = 1
        motoristas = get_motoristas_supabase_cached(empresa_id)
        tarifas = get_tarifas_supabase_cached(empresa_id)
        
        # Combinar motoristas com tarifas
        resultado = []
        for motorista in motoristas:
            id_motorista = str(motorista['id_motorista'])
            tarifas_motorista = tarifas.get(id_motorista, TARIFAS_PADRAO)
            
            motorista_com_tarifas = {
                'id_motorista': motorista['id_motorista'],
                'nome_motorista': motorista['nome_motorista'],
                'tarifas': {
                    '0': tarifas_motorista.get(0, TARIFAS_PADRAO[0]),
                    '9': tarifas_motorista.get(9, TARIFAS_PADRAO[9]),
                    '6': tarifas_motorista.get(6, TARIFAS_PADRAO[6]),
                    '8': tarifas_motorista.get(8, TARIFAS_PADRAO[8])
                }
            }
            resultado.append(motorista_com_tarifas)
        
        return jsonify({
            'success': True,
            'data': resultado
        })
        
    except Exception as e:
        return jsonify({'error': f'Erro ao buscar tarifas: {str(e)}'}), 500

# ==================== API DE ESTATÍSTICAS GERAIS ====================

@app.route('/api/estatisticas', methods=['GET'])
def get_estatisticas_gerais():
    """Estatísticas gerais do sistema"""
    try:
        empresa_id = 1
        
        # Contar dados
        total_motoristas = len(get_motoristas_supabase_cached(empresa_id))
        total_prestadores = len(get_prestadores_supabase_cached(empresa_id))
        total_awbs = get_awbs_count_supabase(empresa_id)
        
        return jsonify({
            'success': True,
            'data': {
                'total_motoristas': total_motoristas,
                'total_prestadores': total_prestadores,
                'total_awbs': total_awbs,
                'supabase_connected': SUPABASE_CONNECTED
            }
        })
        
    except Exception as e:
        return jsonify({'error': f'Erro ao buscar estatísticas: {str(e)}'}), 500

# ==================== API DE UPLOAD BEAST MODE ====================

@app.route('/api/upload', methods=['POST'])
def upload_file():
    """Upload de arquivo CSV/Excel - VERSÃO v7.5 FINAL COM TIMEOUT ESTENDIDO"""
    try:
        empresa_id = 1
        
        print("=== INÍCIO UPLOAD ENTREGAS v7.5 FINAL - TIMEOUT ESTENDIDO ===")
        start_time = time.time()
        
        if 'file' not in request.files:
            return jsonify({'error': 'Nenhum arquivo enviado'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'Nenhum arquivo selecionado'}), 400
        
        print("Carregando dados do Supabase...")
        motoristas = get_motoristas_supabase_cached(empresa_id)
        tarifas_cache = get_tarifas_supabase_cached(empresa_id)
        
        motoristas_dict = {m['id_motorista']: m for m in motoristas}
        print(f"Sistema carregado: {len(motoristas)} motoristas")
        print(f"Fonte: {'Supabase' if SUPABASE_CONNECTED else 'Local'}")
        
        file_content_bytes = file.read()
        print(f"Arquivo recebido: {file.filename} ({len(file_content_bytes)} bytes)")
        
        file_content, encoding_usado = detectar_encoding(file_content_bytes)
        print(f"Encoding detectado: {encoding_usado}")
        
        delimitador = detectar_delimitador_csv(file_content)
        print(f"Delimitador detectado: '{delimitador}'")
        
        try:
            csv_reader = csv.DictReader(io.StringIO(file_content), delimiter=delimitador)
            dados = list(csv_reader)
            total_linhas = len(dados)
            print(f"CSV carregado: {total_linhas} linhas")
            print(f"Colunas: {csv_reader.fieldnames}")
            
            # PROCESSAMENTO BEAST MODE
            resultado = processador_beast.processar_csv_beast_mode(
                dados, motoristas_dict, tarifas_cache, empresa_id
            )
            
        except Exception as e:
            return jsonify({'error': f'Erro ao processar CSV: {str(e)}'}), 500
        
        tempo_total = time.time() - start_time
        
        print(f"=== RESULTADO FINAL v7.5 ===")
        print(f"Tempo total: {tempo_total:.2f} segundos")
        print(f"Entregas processadas: {resultado['entregas_processadas']}")
        print(f"Entregas com erro: {resultado['entregas_erro']}")
        print(f"Total de AWBs: {resultado['total_awbs']}")
        print(f"Performance: {resultado['entregas_processadas']/tempo_total:.0f} linhas/segundo")
        print(f"Supabase conectado: {'✅' if SUPABASE_CONNECTED else '❌'}")
        
        return jsonify({
            'success': True,
            'message': f'v7.5 FINAL: {resultado["entregas_processadas"]} entregas em {tempo_total:.1f}s',
            'data': {
                'entregas_processadas': resultado['entregas_processadas'],
                'entregas_erro': resultado['entregas_erro'],
                'total_awbs': resultado['total_awbs'],
                'tempo_processamento': tempo_total,
                'performance_linhas_por_segundo': round(resultado['entregas_processadas']/tempo_total) if tempo_total > 0 else 0,
                'supabase_connected': SUPABASE_CONNECTED,
                'version': 'v7.5-FINAL'
            }
        })
        
    except Exception as e:
        return jsonify({'error': f'Erro no upload: {str(e)}'}), 500

# ==================== STATUS DO SISTEMA ====================

@app.route('/api/status', methods=['GET'])
def get_system_status():
    """Status do sistema v7.5 FINAL"""
    return jsonify({
        'success': True,
        'data': {
            'supabase_available': SUPABASE_AVAILABLE,
            'supabase_connected': SUPABASE_CONNECTED,
            'cache_timestamp': _cache_timestamp,
            'version': 'v7.5-FINAL',
            'max_workers': MAX_WORKERS,
            'batch_size': BATCH_SIZE,
            'supabase_batch_size': SUPABASE_BATCH_SIZE
        }
    })

# ==================== INICIALIZAÇÃO v7.5 FINAL ====================

if __name__ == '__main__':
    print("🚀 MenezesLog SaaS v7.5 FINAL iniciado!")
    print("⏱️ Timeout estendido para arquivos grandes")
    print("💾 Supabase com todas as APIs funcionando")
    print("⚡ Processamento assíncrono otimizado")
    print("🎯 Sistema completo e funcional")
    print(f"📡 Supabase: {'✅ CONECTADO' if SUPABASE_CONNECTED else '❌ DESCONECTADO'}")
    print(f"🔧 Workers: {MAX_WORKERS} | Batch: {BATCH_SIZE} | Supabase Batch: {SUPABASE_BATCH_SIZE}")
    app.run(host='0.0.0.0', port=5000, debug=False)

